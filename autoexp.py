"""autoexp.py — FP32 / QAT / QAT+Snapshot Ensemble(top-5) / PTQ 비교 실험.

실행:
    python autoexp.py

결과:
    ./results/autoexp/comparison_table.txt  (콘솔 출력과 동일)
    ./results/autoexp/comparison.xlsx
"""

import os
os.environ['TF_USE_LEGACY_KERAS'] = '1'
import warnings
import logging

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
warnings.filterwarnings('ignore')
logging.getLogger('tensorflow').setLevel(logging.ERROR)
logging.getLogger('absl').setLevel(logging.ERROR)

import numpy as np
import tensorflow as tf
tf.get_logger().setLevel('ERROR')

import config
from dataloader  import load_raw_data
from batchloader import get_batches
from model       import build_model
from modules     import CATNet, ChannelAttention
from metrics     import compute_metrics
from loss        import get_loss, get_optimizer

from easyquant import (
    LayerRule, QuantizerSpec,
    QATBuilder, PTQBuilder,
    CosineRestartSchedule, LRSchedulerCallback, SnapshotSaver,
)

# ─────────────────────────────────────────────────────────────────────────────
# Experiment hyper-parameters  (여기만 수정)
# ─────────────────────────────────────────────────────────────────────────────

EXP_DIR        = './results/autoexp'
SNAPSHOT_DIR   = os.path.join(EXP_DIR, 'snapshots')
FP32_WEIGHTS   = os.path.join(EXP_DIR, 'fp32.weights.h5')

QAT_CYCLE_LEN  = 10        # cosine restart 주기 (에포크)
QAT_N_CYCLES   = 4         # 총 사이클 수  → epochs = QAT_CYCLE_LEN * QAT_N_CYCLES
QAT_LR_MAX     = 1e-3
QAT_LR_MIN     = 1e-5
TOP_K          = 5         # snapshot ensemble 개수

PTQ_REPR_SIZE  = 256       # calibration 샘플 수

# ─────────────────────────────────────────────────────────────────────────────
# Quantization config (QAT/PTQ 공유)
# ─────────────────────────────────────────────────────────────────────────────

layers = tf.keras.layers

QAT_RULES = [
    LayerRule(
        target_types=[layers.Conv1D, layers.Conv2D],
        weight_quantizer=QuantizerSpec("learnable_scale",     {"num_bits": 8}),
        activation_quantizer=QuantizerSpec("learnable_threshold", {"num_bits": 4}),
    ),
    LayerRule(
        target_types=[layers.Dense],
        weight_quantizer=QuantizerSpec("learnable_scale", {"num_bits": 8}),
        activation_quantizer=QuantizerSpec("learnable_scale", {"num_bits": 8}),
    ),
    LayerRule(
        name_contains=["softmax"],
        skip=True,
    ),
]

QAT_DEFAULT_RULE = LayerRule(
    weight_quantizer=QuantizerSpec("learnable_scale", {"num_bits": 8}),
    activation_quantizer=QuantizerSpec("learnable_scale", {"num_bits": 8}),
)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_dirs():
    os.makedirs(EXP_DIR,      exist_ok=True)
    os.makedirs(SNAPSHOT_DIR, exist_ok=True)


def _val_split(X_train, y_train, val_ratio=0.1):
    n = int(len(X_train) * val_ratio)
    return X_train[:n], y_train[:n], X_train[n:], y_train[n:]


def _eval_keras(model, X_test, y_test):
    """Keras 모델 → (acc, macro_f1)."""
    y_proba = model.predict(X_test, verbose=0)
    y_pred  = np.argmax(y_proba, axis=-1)
    m = compute_metrics(y_test.astype(int), y_pred, y_proba)
    return m['acc'], m['macro_f1'], y_proba


def _print_table(rows):
    """rows: list of (label, acc, f1) tuples."""
    header = f"{'Method':<30} {'Acc':>8} {'Macro F1':>10}"
    sep    = '-' * len(header)
    lines  = [sep, header, sep]
    for label, acc, f1 in rows:
        lines.append(f"{label:<30} {acc*100:>7.2f}%  {f1*100:>8.2f}%")
    lines.append(sep)
    out = '\n'.join(lines)
    print(out)
    return out


def _save_table(text):
    path = os.path.join(EXP_DIR, 'comparison_table.txt')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(text)
    print(f'\nTable saved → {path}')


def _save_excel(rows):
    """간단한 비교 Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print('[WARN] openpyxl 없음. Excel 저장 건너뜀.')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparison'

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    center      = Alignment(horizontal='center')

    headers = ['Method', 'Accuracy (%)', 'Macro F1 (%)']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center

    for ri, (label, acc, f1) in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=label)
        ws.cell(row=ri, column=2, value=round(acc * 100, 2))
        ws.cell(row=ri, column=3, value=round(f1  * 100, 2))

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16

    path = os.path.join(EXP_DIR, 'comparison.xlsx')
    wb.save(path)
    print(f'Excel  saved → {path}')

# ─────────────────────────────────────────────────────────────────────────────
# Step 1: FP32 baseline
# ─────────────────────────────────────────────────────────────────────────────

def run_fp32(X_train, y_train, X_test, y_test):
    print('\n[1/4] FP32 baseline')
    model = build_model()

    n_val = int(len(X_train) * 0.1)
    Xv, yv = X_train[:n_val], y_train[:n_val]
    Xt, yt = X_train[n_val:], y_train[n_val:]

    model.compile(
        optimizer=get_optimizer(),
        loss=get_loss(),
        metrics=['accuracy'],
    )
    model.fit(
        Xt, yt,
        validation_data=(Xv, yv),
        epochs=config.EPOCHS,
        batch_size=config.BATCH_SIZE,
        verbose=1,
    )
    model.save_weights(FP32_WEIGHTS)
    acc, f1, _ = _eval_keras(model, X_test, y_test)
    print(f'  FP32  acc={acc*100:.2f}%  f1={f1*100:.2f}%')
    return model, acc, f1

# ─────────────────────────────────────────────────────────────────────────────
# Step 2: QAT (single model, last epoch)
# ─────────────────────────────────────────────────────────────────────────────

def run_qat(fp_model, X_train, y_train, X_test, y_test):
    print('\n[2/4] QAT training')

    n_val = int(len(X_train) * 0.1)
    Xv, yv = X_train[:n_val], y_train[:n_val]
    Xt, yt = X_train[n_val:], y_train[n_val:]

    qat_model = QATBuilder(QAT_RULES, QAT_DEFAULT_RULE).build(fp_model)
    qat_model.compile(
        optimizer=tf.keras.optimizers.Adam(QAT_LR_MAX),
        loss=get_loss(),
        metrics=['accuracy'],
    )

    schedule = CosineRestartSchedule(
        lr_max=QAT_LR_MAX, lr_min=QAT_LR_MIN, cycle_length=QAT_CYCLE_LEN
    )
    qat_model.fit(
        Xt, yt,
        validation_data=(Xv, yv),
        epochs=QAT_CYCLE_LEN * QAT_N_CYCLES,
        batch_size=config.BATCH_SIZE,
        callbacks=[
            LRSchedulerCallback(schedule),
            SnapshotSaver(SNAPSHOT_DIR, cycle_length=QAT_CYCLE_LEN),
        ],
        verbose=1,
    )

    acc, f1, _ = _eval_keras(qat_model, X_test, y_test)
    print(f'  QAT  acc={acc*100:.2f}%  f1={f1*100:.2f}%')
    return qat_model, acc, f1

# ─────────────────────────────────────────────────────────────────────────────
# Step 3: QAT + Snapshot Ensemble (top-K by val accuracy)
# ─────────────────────────────────────────────────────────────────────────────

def run_snapshot_ensemble(fp_model, X_train, y_train, X_test, y_test):
    print(f'\n[3/4] QAT snapshot ensemble (top-{TOP_K})')

    n_val = int(len(X_train) * 0.1)
    Xv, yv = X_train[:n_val], y_train[:n_val]

    # 저장된 snapshot 파일 목록
    snap_files = sorted([
        os.path.join(SNAPSHOT_DIR, f)
        for f in os.listdir(SNAPSHOT_DIR)
        if f.startswith('snapshot_') and f.endswith('.weights.h5')
    ])

    if not snap_files:
        print('  [WARN] snapshot 파일 없음. QAT 단계를 먼저 실행하세요.')
        return None, 0.0, 0.0

    # 각 snapshot의 val accuracy 평가
    print(f'  Found {len(snap_files)} snapshots. Evaluating val accuracy...')
    snap_scores = []
    for path in snap_files:
        qat_m = QATBuilder(QAT_RULES, QAT_DEFAULT_RULE).build(fp_model)
        qat_m.compile(
            optimizer=tf.keras.optimizers.Adam(QAT_LR_MAX),
            loss=get_loss(),
            metrics=['accuracy'],
        )
        qat_m.load_weights(path)
        val_proba = qat_m.predict(Xv, verbose=0)
        val_pred  = np.argmax(val_proba, axis=-1)
        val_acc   = np.mean(val_pred == yv)
        snap_scores.append((val_acc, path))
        print(f'    {os.path.basename(path)}: val_acc={val_acc*100:.2f}%')

    # top-K 선택
    snap_scores.sort(key=lambda x: x[0], reverse=True)
    top_k = snap_scores[:TOP_K]
    print(f'  Top-{TOP_K}: {[os.path.basename(p) for _, p in top_k]}')

    # Ensemble: 각 모델의 softmax 평균
    ensemble_proba = np.zeros((len(X_test), config.N_CLASSES), dtype=np.float32)
    for _, path in top_k:
        qat_m = QATBuilder(QAT_RULES, QAT_DEFAULT_RULE).build(fp_model)
        qat_m.compile(
            optimizer=tf.keras.optimizers.Adam(QAT_LR_MAX),
            loss=get_loss(),
        )
        qat_m.load_weights(path)
        ensemble_proba += qat_m.predict(X_test, verbose=0)
    ensemble_proba /= len(top_k)

    y_pred = np.argmax(ensemble_proba, axis=-1)
    m = compute_metrics(y_test.astype(int), y_pred, ensemble_proba)
    acc, f1 = m['acc'], m['macro_f1']
    print(f'  Ensemble acc={acc*100:.2f}%  f1={f1*100:.2f}%')
    return ensemble_proba, acc, f1

# ─────────────────────────────────────────────────────────────────────────────
# Step 4: PTQ (int8)
# ─────────────────────────────────────────────────────────────────────────────

def run_ptq(fp_model, X_train, X_test, y_test):
    print('\n[4/4] PTQ int8')

    # calibration 데이터 (train 앞부분 사용)
    calib = X_train[:PTQ_REPR_SIZE].astype(np.float32)

    def repr_gen():
        for x in calib:
            yield [x[np.newaxis]]

    ptq_builder = PTQBuilder(QAT_RULES, QAT_DEFAULT_RULE)
    tflite_bytes = ptq_builder.build(fp_model, representative_dataset=repr_gen, mode="int8")

    # TFLite 파일 저장
    tflite_path = os.path.join(EXP_DIR, 'model_int8.tflite')
    with open(tflite_path, 'wb') as f:
        f.write(tflite_bytes)
    print(f'  TFLite saved → {tflite_path}')

    # 평가: (x, y) 쌍 iterable 형태로 전달
    test_dataset = zip(X_test, y_test)
    acc = ptq_builder.evaluate(tflite_bytes, test_dataset)

    # f1 계산을 위해 따로 추론
    interpreter = tf.lite.Interpreter(model_content=tflite_bytes)
    interpreter.allocate_tensors()
    in_det  = interpreter.get_input_details()[0]
    out_det = interpreter.get_output_details()[0]
    in_scale  = in_det['quantization'][0]  or 1.0
    in_zp     = in_det['quantization'][1]
    out_scale = out_det['quantization'][0] or 1.0
    out_zp    = out_det['quantization'][1]

    probas = []
    for x in X_test:
        x_in = x[np.newaxis].astype(np.float32)
        if in_det['dtype'] == np.int8:
            x_in = np.round(x_in / in_scale + in_zp).astype(np.int8)
        interpreter.set_tensor(in_det['index'], x_in)
        interpreter.invoke()
        out = interpreter.get_tensor(out_det['index'])
        if out_det['dtype'] == np.int8:
            out = (out.astype(np.float32) - out_zp) * out_scale
        probas.append(out[0])
    probas = np.array(probas)

    y_pred = np.argmax(probas, axis=-1)
    m  = compute_metrics(y_test.astype(int), y_pred, probas)
    f1 = m['macro_f1']
    print(f'  PTQ  acc={acc*100:.2f}%  f1={f1*100:.2f}%')
    return acc, f1

# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    _ensure_dirs()

    print('Loading data...')
    X, Y = load_raw_data()
    X_train, X_test, y_train, y_test = get_batches(X, Y)
    print(f'  Train: {X_train.shape}  Test: {X_test.shape}')

    results = []

    # 1. FP32
    fp_model, fp_acc, fp_f1 = run_fp32(X_train, y_train, X_test, y_test)
    results.append(('FP32', fp_acc, fp_f1))

    # 2. QAT
    _, qat_acc, qat_f1 = run_qat(fp_model, X_train, y_train, X_test, y_test)
    results.append(('QAT', qat_acc, qat_f1))

    # 3. QAT + Snapshot Ensemble (top-5)
    _, ens_acc, ens_f1 = run_snapshot_ensemble(fp_model, X_train, y_train, X_test, y_test)
    results.append((f'QAT + Snapshot Ensemble (top-{TOP_K})', ens_acc, ens_f1))

    # 4. PTQ int8
    ptq_acc, ptq_f1 = run_ptq(fp_model, X_train, X_test, y_test)
    results.append(('PTQ (int8)', ptq_acc, ptq_f1))

    # ── 비교 테이블 출력 및 저장 ──────────────────────────────────────────
    print('\n\n' + '=' * 55)
    print('  COMPARISON RESULTS')
    print('=' * 55)
    table_text = _print_table(results)
    _save_table(table_text)
    _save_excel(results)


if __name__ == '__main__':
    main()
