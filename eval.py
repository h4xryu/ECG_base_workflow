import os

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import seaborn
import tensorflow as tf
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import confusion_matrix
from sklearn.manifold import TSNE

import config
from metrics import compute_metrics


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _make_exp_dir(exp_name: str) -> str:
    path = os.path.join(config.RESULTS_DIR, exp_name)
    os.makedirs(path, exist_ok=True)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# t-SNE
# ─────────────────────────────────────────────────────────────────────────────

def _extract_features(model, X: np.ndarray) -> np.ndarray:
    """Return Flatten-layer output; fall back to raw signal if not found."""
    flat_outputs = [l.output for l in model.layers
                    if isinstance(l, tf.keras.layers.Flatten)]
    if not flat_outputs:
        return X.squeeze(-1)
    feat_model = tf.keras.Model(inputs=model.inputs, outputs=flat_outputs[0])
    return feat_model.predict(X, verbose=0)


def save_tsne_data(model, X_tsne: np.ndarray, y_true_tsne: np.ndarray,
                   exp_dir: str):
    """
    Self-contained t-SNE helper.

    X_tsne      : (n, L, 1) numpy array aligned with y_true_tsne
    y_true_tsne : (n,) class indices OR (n, C) multi-label int array
    """
    print('Running t-SNE…')
    n   = min(len(X_tsne), config.TSNE_MAX_SAMPLES)
    idx = np.random.choice(len(X_tsne), n, replace=False)

    X_s     = X_tsne[idx]
    y_s     = y_true_tsne[idx]
    y_proba = model.predict(X_s, batch_size=config.BATCH_SIZE, verbose=0)

    if config.MULTI_LABEL:
        y_pred = (y_proba > 0.5).astype(np.int32)
    else:
        y_pred = np.argmax(y_proba, axis=-1).astype(np.int32)

    feats = _extract_features(model, X_s)
    emb   = TSNE(n_components=2, perplexity=config.TSNE_PERPLEXITY,
                 random_state=42, n_jobs=-1).fit_transform(feats)

    np.savez(
        os.path.join(exp_dir, 'tsne_data.npz'),
        embeddings   = emb,
        labels       = y_s.astype(np.int32),
        predictions  = y_pred,
        probabilities= y_proba.astype(np.float32),
        samples      = X_s.squeeze(-1).astype(np.float32),
    )
    print(f't-SNE saved → {exp_dir}/tsne_data.npz')


# ─────────────────────────────────────────────────────────────────────────────
# Plots (saved to disk)
# ─────────────────────────────────────────────────────────────────────────────

def plot_history(history, exp_dir: str):
    if history is None:
        return
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))
    ax1.set_title('Accuracy'); ax1.plot(history.history['accuracy'],     label='train')
    ax1.plot(history.history['val_accuracy'], label='val'); ax1.legend()
    ax2.set_title('Loss');     ax2.plot(history.history['loss'],         label='train')
    ax2.plot(history.history['val_loss'],     label='val'); ax2.legend()
    plt.tight_layout()
    fig.savefig(os.path.join(exp_dir, 'training_curves.png'), dpi=150)
    plt.close(fig)


def plot_confusion_matrix(y_true, y_pred, exp_dir: str) -> np.ndarray:
    cm = confusion_matrix(y_true, y_pred)
    fig, ax = plt.subplots(figsize=(7, 6))
    seaborn.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=config.CLASS_NAMES,
                    yticklabels=config.CLASS_NAMES, ax=ax)
    ax.set_xlabel('Predicted'); ax.set_ylabel('True')
    plt.tight_layout()
    fig.savefig(os.path.join(exp_dir, 'confusion_matrix.png'), dpi=150)
    plt.close(fig)
    return cm


def plot_confusion_matrix_multilabel(y_true, y_pred, exp_dir: str) -> list:
    """Per-label binary confusion matrices for multi-label classification."""
    n  = config.N_CLASSES
    nc = min(n, 3)
    nr = (n + nc - 1) // nc
    fig, axes = plt.subplots(nr, nc, figsize=(nc * 4, nr * 3.5))
    axes = np.array(axes).reshape(-1)
    cms = []
    for i, name in enumerate(config.CLASS_NAMES):
        cm_i = confusion_matrix(y_true[:, i], y_pred[:, i], labels=[0, 1])
        cms.append(cm_i)
        seaborn.heatmap(cm_i, annot=True, fmt='d', cmap='Blues',
                        xticklabels=['Neg', 'Pos'],
                        yticklabels=['Neg', 'Pos'], ax=axes[i])
        axes[i].set_title(name, fontsize=9)
        axes[i].set_xlabel('Predicted'); axes[i].set_ylabel('True')
    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)
    plt.tight_layout()
    fig.savefig(os.path.join(exp_dir, 'confusion_matrix.png'), dpi=150)
    plt.close(fig)
    return cms


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

_H_FILL  = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
_SH_FILL = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
_M_FILL  = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
_H_FONT  = Font(bold=True, color='FFFFFF', size=11)
_SH_FONT = Font(bold=True, size=10)
_CTR     = Alignment(horizontal='center', vertical='center')
_BDR     = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))


def _c(ws, row, col, value=None, fill=None, font=None):
    """Set cell value + style shorthand."""
    cell = ws.cell(row=row, column=col)
    if value is not None: cell.value = value
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    cell.alignment = _CTR
    cell.border    = _BDR
    return cell


def _create_performance_sheet(wb, metrics: dict, exp_name: str):
    ws      = wb.create_sheet('Performance Metrics', 0)
    classes = config.CLASS_NAMES
    mlbls   = ['Acc', 'Se', 'Sp', 'Pr', 'F1']

    # ── Row 1: group headers ──────────────────────────────────────────────
    _c(ws, 1, 1, 'Experiment', _H_FILL, _H_FONT)

    ws.merge_cells('B1:F1');  _c(ws, 1, 2,  'Macro',     _H_FILL, _H_FONT)
    ws.merge_cells('G1:K1');  _c(ws, 1, 7,  'Weighted',  _H_FILL, _H_FONT)

    per_end = 11 + len(classes) * 5
    ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=per_end)
    _c(ws, 1, 12, 'Per-Class', _H_FILL, _H_FONT)

    # ── Row 2: sub-headers ────────────────────────────────────────────────
    _c(ws, 2, 1, 'Name', _SH_FILL, _SH_FONT)
    for i, m in enumerate(mlbls):
        _c(ws, 2, 2 + i, m, _SH_FILL, _SH_FONT)   # Macro
        _c(ws, 2, 7 + i, m, _SH_FILL, _SH_FONT)   # Weighted

    for ci, cls in enumerate(classes):
        sc = 12 + ci * 5
        ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=sc + 4)
        _c(ws, 2, sc, cls, _SH_FILL, _SH_FONT)

    # ── Row 3: per-class metric names ─────────────────────────────────────
    ws.cell(row=3, column=1).value = ''
    for ci in range(len(classes)):
        for j, m in enumerate(mlbls):
            _c(ws, 3, 12 + ci * 5 + j, m, _M_FILL, Font(size=9))

    # ── Row 4: data ───────────────────────────────────────────────────────
    r = 4
    _c(ws, r, 1, exp_name)

    macro_vals = [metrics['acc'],        metrics['macro_recall'],
                  metrics['macro_specificity'], metrics['macro_precision'], metrics['macro_f1']]
    w_vals     = [metrics['acc'],        metrics['w_recall'],
                  metrics['w_specificity'],     metrics['w_precision'],     metrics['w_f1']]

    for i, v in enumerate(macro_vals):
        c = _c(ws, r, 2 + i, round(v * 100, 2))
        c.number_format = '0.00'
    for i, v in enumerate(w_vals):
        c = _c(ws, r, 7 + i, round(v * 100, 2))
        c.number_format = '0.00'
    for ci in range(len(classes)):
        per = [metrics['pc_acc'][ci], metrics['pc_se'][ci], metrics['pc_sp'][ci],
               metrics['pc_pr'][ci],  metrics['pc_f1'][ci]]
        for j, v in enumerate(per):
            c = _c(ws, r, 12 + ci * 5 + j, round(v * 100, 2))
            c.number_format = '0.00'

    # Column widths
    ws.column_dimensions['A'].width = 35
    for col in range(2, per_end + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9


def _create_confusion_matrix_sheet(wb, cm: np.ndarray):
    ws      = wb.create_sheet('Confusion Matrix')
    classes = config.CLASS_NAMES

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    _c(ws, 1, 1, 'Confusion Matrix', _H_FILL, _H_FONT)

    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=7)
    _c(ws, 2, 3, 'Predicted', _SH_FILL, _SH_FONT)

    for i, cls in enumerate(classes):
        _c(ws, 3, 3 + i, cls, _SH_FILL, _SH_FONT)

    ws.merge_cells(start_row=4, start_column=1, end_row=8, end_column=1)
    cell = ws.cell(row=4, column=1)
    cell.value = 'Actual'
    cell.fill  = _SH_FILL; cell.font = _SH_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
    cell.border    = _BDR

    for i, cls in enumerate(classes):
        _c(ws, 4 + i, 2, cls, _SH_FILL, _SH_FONT)
        for j in range(len(classes)):
            c = _c(ws, 4 + i, 3 + j, int(cm[i, j]))
            c.number_format = '0'

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _create_performance_sheet_multilabel(wb, metrics: dict, exp_name: str):
    ws      = wb.create_sheet('Performance Metrics', 0)
    classes = config.CLASS_NAMES

    headers = ['Experiment', 'Subset Acc', 'Hamming Loss',
               'Macro Pr', 'Macro Re', 'Macro F1',
               'Micro Pr', 'Micro Re', 'Micro F1',
               'W Pr',     'W Re',     'W F1']
    for ci, h in enumerate(headers, 1):
        _c(ws, 1, ci, h, _H_FILL, _H_FONT)

    col = len(headers) + 1
    for name in classes:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        _c(ws, 1, col, name, _H_FILL, _H_FONT)
        for j, sub in enumerate(['Pr', 'Re', 'F1', 'AUC']):
            _c(ws, 2, col + j, sub, _SH_FILL, _SH_FONT)
        col += 4

    r = 3
    _c(ws, r, 1, exp_name)
    summary = [
        metrics['subset_accuracy'], metrics['hamming_loss'],
        metrics['macro_precision'],  metrics['macro_recall'],  metrics['macro_f1'],
        metrics['micro_precision'],  metrics['micro_recall'],  metrics['micro_f1'],
        metrics['w_precision'],      metrics['w_recall'],      metrics['w_f1'],
    ]
    for ci, v in enumerate(summary, 2):
        _c(ws, r, ci, round(v * 100, 2)).number_format = '0.00'

    col = len(headers) + 1
    for i in range(len(classes)):
        auc = metrics['per_label_auc'][i] if metrics['per_label_auc'] else 0.0
        for j, v in enumerate([metrics['pc_precision'][i], metrics['pc_recall'][i],
                                metrics['pc_f1'][i], auc]):
            _c(ws, r, col + j, round(v * 100, 2)).number_format = '0.00'
        col += 4

    ws.column_dimensions['A'].width = 35
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 10


def save_excel(metrics: dict, cm, exp_name: str, exp_dir: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if config.MULTI_LABEL:
        _create_performance_sheet_multilabel(wb, metrics, exp_name)
    else:
        _create_performance_sheet(wb, metrics, exp_name)
        _create_confusion_matrix_sheet(wb, cm)
    path = os.path.join(exp_dir, f'{exp_name}.xlsx')
    wb.save(path)
    print(f'Excel  saved → {path}')


# ─────────────────────────────────────────────────────────────────────────────
# Model weights
# ─────────────────────────────────────────────────────────────────────────────

def save_weights(model, exp_dir: str):
    path = os.path.join(exp_dir, 'model.weights.h5')
    model.save_weights(path)
    print(f'Weights saved → {path}')


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────────

def full_eval(model, X_train, y_train, X_test, y_test,
              history=None, exp_name: str = None,
              X_tsne: np.ndarray = None, y_tsne: np.ndarray = None,
              y_proba: np.ndarray = None) -> dict:
    """
    X_test   : numpy array, tf.data.Dataset, or None (when y_proba is given)
    y_proba  : pre-computed probabilities (N, C).  When provided, X_test is
               not used for prediction — pass this to bypass model.predict
               and avoid the cudnn DEVICE_TYPE_INVALID error on lazy datasets.
    X_tsne   : small numpy array (TSNE_MAX_SAMPLES rows) aligned with y_tsne.
    y_tsne   : labels aligned with X_tsne.
    """
    if exp_name is None:
        exp_name = config.get_exp_name()
    exp_dir = _make_exp_dir(exp_name)
    print(f'\n── Evaluating: {exp_name} ──')

    if y_proba is None:
        is_dataset = isinstance(X_test, tf.data.Dataset)
        if is_dataset:
            y_proba = model.predict(X_test, verbose=0)
        else:
            X_test  = np.nan_to_num(X_test, nan=0.0, posinf=0.0, neginf=0.0)
            y_proba = model.predict(X_test, batch_size=config.BATCH_SIZE, verbose=0)

    if config.MULTI_LABEL:
        y_true  = (y_test > 0).astype(int)
        y_pred  = (y_proba > 0.5).astype(int)
        metrics = compute_metrics(y_true, y_pred, y_proba)
        cm      = plot_confusion_matrix_multilabel(y_true, y_pred, exp_dir)

        plot_history(history, exp_dir)
        save_excel(metrics, cm, exp_name, exp_dir)
        save_weights(model, exp_dir)

        if config.TSNE_ENABLED:
            _X = X_tsne if X_tsne is not None else (None if is_dataset else X_test)
            _y = (y_tsne > 0).astype(int) if y_tsne is not None else (
                 y_true if not is_dataset else None)
            if _X is not None and _y is not None:
                save_tsne_data(model, _X, _y, exp_dir)
            else:
                print('[eval] Skipping t-SNE: X_tsne/y_tsne not provided.')

        print(f"  Subset Acc  : {metrics['subset_accuracy']*100:.2f}%")
        print(f"  Hamming Loss: {metrics['hamming_loss']*100:.4f}")
        print(f"  Macro  F1   : {metrics['macro_f1']*100:.2f}%")
        for i, name in enumerate(config.CLASS_NAMES):
            print(f"  {name}  Pr={metrics['pc_precision'][i]*100:.1f}% "
                  f"Re={metrics['pc_recall'][i]*100:.1f}% "
                  f"F1={metrics['pc_f1'][i]*100:.1f}%")
    else:
        y_pred  = np.argmax(y_proba, axis=-1)
        metrics = compute_metrics(y_test.astype(int), y_pred, y_proba)
        cm      = plot_confusion_matrix(y_test, y_pred, exp_dir)

        plot_history(history, exp_dir)
        save_excel(metrics, cm, exp_name, exp_dir)
        save_weights(model, exp_dir)

        if config.TSNE_ENABLED:
            _X = X_tsne if X_tsne is not None else (None if is_dataset else X_test)
            _y = y_tsne.astype(int) if y_tsne is not None else (
                 y_test.astype(int) if not is_dataset else None)
            if _X is not None and _y is not None:
                save_tsne_data(model, _X, _y, exp_dir)
            else:
                print('[eval] Skipping t-SNE: X_tsne/y_tsne not provided.')

        print(f"  Overall Acc : {metrics['acc']*100:.2f}%")
        print(f"  Macro  F1   : {metrics['macro_f1']*100:.2f}%")
        for i, name in enumerate(config.CLASS_NAMES):
            print(f"  {name}  Se={metrics['pc_se'][i]*100:.1f}% "
                  f"Sp={metrics['pc_sp'][i]*100:.1f}% "
                  f"F1={metrics['pc_f1'][i]*100:.1f}%")

    return metrics


def full_eval_hicardi(model, X_mmap, Y, te_idx, history=None, exp_name=None):
    """Hicardi 전용 래퍼 — mmap에서 직접 추론·t-SNE 준비 후 full_eval 호출."""
    from batchloader_hicardi import predict_from_mmap, TARGET_LENGTH

    y_test  = Y[te_idx]
    print(f'[eval] Running chunk-based predict on {len(te_idx):,} val samples…')
    y_proba = predict_from_mmap(model, X_mmap, te_idx)

    tsne_n  = min(config.TSNE_MAX_SAMPLES, len(te_idx))
    X_tsne  = np.array(X_mmap[te_idx[:tsne_n]], dtype=np.float32).reshape(-1, TARGET_LENGTH, 1)
    np.nan_to_num(X_tsne, copy=False, nan=0.0, posinf=0.0, neginf=0.0)

    return full_eval(model, None, None, None, y_test, history, exp_name,
                     X_tsne=X_tsne, y_tsne=y_test[:tsne_n], y_proba=y_proba)
